#!/usr/bin/env python3
"""
Create a properly formatted LibreOffice spreadsheet with merged cells
for the tap and drill sizes table.

Data is structured to minimize duplication:
- Drill sizes are defined once in lookup tables
- Thread data is organized hierarchically by screw size
- Clearance drill specifications are shared across thread pitches for each screw size
"""

from fractions import Fraction
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# ============================================================================
# DRILL SIZE LOOKUP TABLES (from Wikipedia: Drill bit sizes)
# ============================================================================

# Number drill sizes (80 smallest to 1 largest)
# Source: https://en.wikipedia.org/wiki/Drill_bit_sizes
NUMBER_DRILLS = {
    80: 0.0135, 79: 0.0145, 78: 0.0160, 77: 0.0180, 76: 0.0200,
    75: 0.0210, 74: 0.0225, 73: 0.0240, 72: 0.0250, 71: 0.0260,
    70: 0.0280, 69: 0.0292, 68: 0.0310, 67: 0.0320, 66: 0.0330,
    65: 0.0350, 64: 0.0360, 63: 0.0370, 62: 0.0380, 61: 0.0390,
    60: 0.0400, 59: 0.0410, 58: 0.0420, 57: 0.0430, 56: 0.0465,
    55: 0.0520, 54: 0.0550, 53: 0.0595, 52: 0.0635, 51: 0.0670,
    50: 0.0700, 49: 0.0730, 48: 0.0760, 47: 0.0785, 46: 0.0810,
    45: 0.0820, 44: 0.0860, 43: 0.0890, 42: 0.0935, 41: 0.0960,
    40: 0.0980, 39: 0.0995, 38: 0.1015, 37: 0.1040, 36: 0.1065,
    35: 0.1100, 34: 0.1110, 33: 0.1130, 32: 0.1160, 31: 0.1200,
    30: 0.1285, 29: 0.1360, 28: 0.1405, 27: 0.1440, 26: 0.1470,
    25: 0.1495, 24: 0.1520, 23: 0.1540, 22: 0.1570, 21: 0.1590,
    20: 0.1610, 19: 0.1660, 18: 0.1695, 17: 0.1730, 16: 0.1770,
    15: 0.1800, 14: 0.1820, 13: 0.1850, 12: 0.1890, 11: 0.1910,
    10: 0.1935, 9: 0.1960, 8: 0.1990, 7: 0.2010, 6: 0.2040,
    5: 0.2055, 4: 0.2090, 3: 0.2130, 2: 0.2210, 1: 0.2280,
}

# Letter drill sizes (A smallest to Z largest)
LETTER_DRILLS = {
    'A': 0.2340, 'B': 0.2380, 'C': 0.2420, 'D': 0.2460, 'E': 0.2500,
    'F': 0.2570, 'G': 0.2610, 'H': 0.2660, 'I': 0.2720, 'J': 0.2770,
    'K': 0.2810, 'L': 0.2900, 'M': 0.2950, 'N': 0.3020, 'O': 0.3160,
    'P': 0.3230, 'Q': 0.3320, 'R': 0.3390, 'S': 0.3480, 'T': 0.3580,
    'U': 0.3680, 'V': 0.3770, 'W': 0.3860, 'X': 0.3970, 'Y': 0.4040,
    'Z': 0.4130,
}

# Fractional drill sizes (in 64ths of an inch, up to 1 inch, plus common larger sizes)
FRACTIONAL_DRILLS = {
    Fraction(1, 64): 0.0156, Fraction(1, 32): 0.0313, Fraction(3, 64): 0.0469,
    Fraction(1, 16): 0.0625, Fraction(5, 64): 0.0781, Fraction(3, 32): 0.0938,
    Fraction(7, 64): 0.1094, Fraction(1, 8): 0.1250, Fraction(9, 64): 0.1406,
    Fraction(5, 32): 0.1563, Fraction(11, 64): 0.1719, Fraction(3, 16): 0.1875,
    Fraction(13, 64): 0.2031, Fraction(7, 32): 0.2188, Fraction(15, 64): 0.2344,
    Fraction(1, 4): 0.2500, Fraction(17, 64): 0.2656, Fraction(9, 32): 0.2812,
    Fraction(19, 64): 0.2969, Fraction(5, 16): 0.3125, Fraction(21, 64): 0.3281,
    Fraction(11, 32): 0.3438, Fraction(23, 64): 0.3594, Fraction(3, 8): 0.3750,
    Fraction(25, 64): 0.3906, Fraction(13, 32): 0.4062, Fraction(27, 64): 0.4219,
    Fraction(7, 16): 0.4375, Fraction(29, 64): 0.4531, Fraction(15, 32): 0.4688,
    Fraction(31, 64): 0.4844, Fraction(1, 2): 0.5000, Fraction(33, 64): 0.5156,
    Fraction(17, 32): 0.5312, Fraction(35, 64): 0.5469, Fraction(9, 16): 0.5625,
    Fraction(37, 64): 0.5781, Fraction(19, 32): 0.5938, Fraction(39, 64): 0.6094,
    Fraction(5, 8): 0.6250, Fraction(41, 64): 0.6406, Fraction(21, 32): 0.6562,
    Fraction(43, 64): 0.6719, Fraction(11, 16): 0.6875, Fraction(45, 64): 0.7031,
    Fraction(23, 32): 0.7188, Fraction(47, 64): 0.7344, Fraction(3, 4): 0.7500,
    Fraction(49, 64): 0.7656, Fraction(25, 32): 0.7812, Fraction(51, 64): 0.7969,
    Fraction(13, 16): 0.8125, Fraction(53, 64): 0.8281, Fraction(27, 32): 0.8438,
    Fraction(55, 64): 0.8594, Fraction(7, 8): 0.8750, Fraction(57, 64): 0.8906,
    Fraction(29, 32): 0.9062, Fraction(59, 64): 0.9219, Fraction(15, 16): 0.9375,
    Fraction(61, 64): 0.9531, Fraction(31, 32): 0.9688, Fraction(63, 64): 0.9844,
    Fraction(1, 1): 1.0000,
    # Larger common sizes over 1 inch (represented as improper fractions)
    Fraction(65, 64): 1.0156,  # 1-1/64 = 65/64
    Fraction(33, 32): 1.0313,  # 1-1/32 = 33/32
}

def get_drill_decimal(drill_spec):
    """
    Look up the decimal equivalent of a drill size.
    
    Args:
        drill_spec: Can be int (number drill), str (letter drill), 
                   Fraction (fractional drill), or float (direct decimal)
    
    Returns:
        float: Decimal equivalent in inches
    """
    if isinstance(drill_spec, int):
        return NUMBER_DRILLS[drill_spec]
    elif isinstance(drill_spec, str):
        return LETTER_DRILLS[drill_spec]
    elif isinstance(drill_spec, Fraction):
        return FRACTIONAL_DRILLS[drill_spec]
    elif isinstance(drill_spec, float):
        return drill_spec
    else:
        raise ValueError(f"Unknown drill spec type: {type(drill_spec)}")

def format_drill_size(drill_spec):
    """
    Format a drill size for display.
    
    Args:
        drill_spec: Can be int (number drill), str (letter drill), 
                   Fraction (fractional drill), or float (direct decimal)
    
    Returns:
        str: Formatted drill size (e.g., "#53", "F", "3/64")
    """
    if isinstance(drill_spec, int):
        return str(drill_spec)
    elif isinstance(drill_spec, str):
        return drill_spec
    elif isinstance(drill_spec, Fraction):
        if drill_spec > 1:
            # Handle mixed fractions like 1-1/64
            whole = drill_spec.numerator // drill_spec.denominator
            remainder = drill_spec.numerator % drill_spec.denominator
            if remainder == 0:
                return str(whole)
            else:
                return f"{whole}-{remainder}/{drill_spec.denominator}"
        elif drill_spec == 1:
            return "1"
        else:
            return f"{drill_spec.numerator}/{drill_spec.denominator}"
    elif isinstance(drill_spec, float):
        return f"{drill_spec:.4f}"
    else:
        raise ValueError(f"Unknown drill spec type: {type(drill_spec)}")

def format_decimal(value):
    """
    Format a decimal value to match the original spreadsheet format.
    All values use 4 decimal places, with values < 1.0 having no leading zero.
    
    Note: The original file has some rounding inconsistencies - 15/32" (0.46875)
    appears as both .4687 and .4688 depending on context.
    """
    formatted = f"{value:.4f}"
    
    if value < 1.0:
        # Remove leading zero for values less than 1 (e.g., '0.0600' -> '.0600')
        formatted = formatted[1:]
    
    # Handle special cases that should have fewer decimals to match original
    # These are the exact values from the original spreadsheet
    if formatted == '.1250':
        return '.125'
    elif formatted == '.1380':
        return '.138'
    elif formatted == '1.0000':
        return '1.000'
    
    return formatted

def format_screw_size(screw_key):
    """
    Format screw size for display in spreadsheet.
    Removes the '#' prefix from numbered screws and the '"' from inch screws.
    
    Examples:
        "#0" -> "0"
        "#12" -> "12"
        "1/4" -> "1/4"
        "1\"" -> "1"
    """
    if screw_key.startswith('#'):
        return screw_key[1:]  # Remove the # prefix
    elif screw_key.endswith('"'):
        return screw_key[:-1]  # Remove the " suffix
    else:
        return screw_key  # Fractional sizes are displayed as-is

# ============================================================================
# THREAD DATA ORGANIZED BY SCREW SIZE
# ============================================================================

# Each screw size contains:
# - major_diameter: Outer diameter of the thread
# - threads: Dict of TPI -> thread specifications
#   - minor_diameter: Inner diameter for this thread pitch
#   - tap_75: Drill for 75% thread (soft materials)
#   - tap_50: Drill for 50% thread (hard materials)
# - clearance: Drill specifications (shared across all TPIs for this screw)
#   - close_fit: Drill for close fit clearance hole
#   - free_fit: Drill for free fit clearance hole

THREAD_DATA = {
    "#0": {
        "major_diameter": 0.0600,
        "threads": {
            80: {
                "minor_diameter": 0.0447,
                "tap_75": Fraction(3, 64),
                "tap_50": 55,
            }
        },
        "clearance": {"close_fit": 52, "free_fit": 50},
        "shcs": {
            "hex": "0.050",
            "counterbore_drill": Fraction(1, 8),
            "counterbore_dia": 0.125,
            "counterbore_depth": 0.074
        }
    },
    "#1": {
        "major_diameter": 0.0730,
        "threads": {
            64: {
                "minor_diameter": 0.0538,
                "tap_75": 53,
                "tap_50": Fraction(1, 16),
            },
            72: {
                "minor_diameter": 0.0560,
                "tap_75": 53,
                "tap_50": 52,
            }
        },
        "clearance": {"close_fit": 48, "free_fit": 46},
        "shcs": {
            "hex": Fraction(1, 16),
            "counterbore_drill": Fraction(5, 32),
            "counterbore_dia": 0.15625,
            "counterbore_depth": 0.087
        }
    },
    "#2": {
        "major_diameter": 0.0860,
        "threads": {
            56: {
                "minor_diameter": 0.0641,
                "tap_75": 50,
                "tap_50": 49,
            },
            64: {
                "minor_diameter": 0.0668,
                "tap_75": 50,
                "tap_50": 48,
            }
        },
        "clearance": {"close_fit": 43, "free_fit": 41},
        "shcs": {
            "hex": Fraction(5, 64),
            "counterbore_drill": Fraction(3, 16),
            "counterbore_dia": 0.1875,
            "counterbore_depth": 0.102
        }
    },
    "#3": {
        "major_diameter": 0.0990,
        "threads": {
            48: {
                "minor_diameter": 0.0734,
                "tap_75": 47,
                "tap_50": 44,
            },
            56: {
                "minor_diameter": 0.0771,
                "tap_75": 45,
                "tap_50": 43,
            }
        },
        "clearance": {"close_fit": 37, "free_fit": 35},
        "shcs": {
            "hex": Fraction(5, 64),
            "counterbore_drill": Fraction(7, 32),
            "counterbore_dia": 0.21875,
            "counterbore_depth": 0.115
        }
    },
    "#4": {
        "major_diameter": 0.1120,
        "threads": {
            40: {
                "minor_diameter": 0.0813,
                "tap_75": 43,
                "tap_50": 41,
            },
            48: {
                "minor_diameter": 0.0864,
                "tap_75": 42,
                "tap_50": 40,
            }
        },
        "clearance": {"close_fit": 32, "free_fit": 30},
        "shcs": {
            "hex": Fraction(3, 32),
            "counterbore_drill": Fraction(7, 32),
            "counterbore_dia": 0.21875,
            "counterbore_depth": 0.130
        }
    },
    "#5": {
        "major_diameter": 0.1250,
        "threads": {
            40: {
                "minor_diameter": 0.0943,
                "tap_75": 38,
                "tap_50": Fraction(7, 64),
            },
            44: {
                "minor_diameter": 0.0971,
                "tap_75": 37,
                "tap_50": 35,
            }
        },
        "clearance": {"close_fit": 30, "free_fit": 29},
        "shcs": {
            "hex": Fraction(3, 32),
            "counterbore_drill": Fraction(1, 4),
            "counterbore_dia": 0.250,
            "counterbore_depth": 0.145
        }
    },
    "#6": {
        "major_diameter": 0.1380,
        "threads": {
            32: {
                "minor_diameter": 0.0997,
                "tap_75": 36,
                "tap_50": 32,
            },
            40: {
                "minor_diameter": 0.1073,
                "tap_75": 33,
                "tap_50": 31,
            }
        },
        "clearance": {"close_fit": 27, "free_fit": 25},
        "shcs": {
            "hex": Fraction(7, 64),
            "counterbore_drill": Fraction(9, 32),
            "counterbore_dia": 0.28125,
            "counterbore_depth": 0.158
        }
    },
    "#8": {
        "major_diameter": 0.1640,
        "threads": {
            32: {
                "minor_diameter": 0.1257,
                "tap_75": 29,
                "tap_50": 27,
            },
            36: {
                "minor_diameter": 0.1299,
                "tap_75": 29,
                "tap_50": 26,
            }
        },
        "clearance": {"close_fit": 18, "free_fit": 16},
        "shcs": {
            "hex": Fraction(9, 64),
            "counterbore_drill": Fraction(5, 16),
            "counterbore_dia": 0.3125,
            "counterbore_depth": 0.188
        }
    },
    "#10": {
        "major_diameter": 0.1900,
        "threads": {
            24: {
                "minor_diameter": 0.1389,
                "tap_75": 25,
                "tap_50": 20,
            },
            32: {
                "minor_diameter": 0.1517,
                "tap_75": 21,
                "tap_50": 18,
            }
        },
        "clearance": {"close_fit": 9, "free_fit": 7},
        "shcs": {
            "hex": Fraction(5, 32),
            "counterbore_drill": Fraction(3, 8),
            "counterbore_dia": 0.375,
            "counterbore_depth": 0.218
        }
    },
    "#12": {
        "major_diameter": 0.2160,
        "threads": {
            24: {
                "minor_diameter": 0.1649,
                "tap_75": 16,
                "tap_50": 12,
            },
            28: {
                "minor_diameter": 0.1722,
                "tap_75": 14,
                "tap_50": 10,
            },
            32: {
                "minor_diameter": 0.1777,
                "tap_75": 13,
                "tap_50": 9,
            }
        },
        "clearance": {"close_fit": 2, "free_fit": 1},
        "shcs": {
            "hex": Fraction(5, 32),
            "counterbore_drill": Fraction(3, 8),
            "counterbore_dia": 0.375,
            "counterbore_depth": 0.218
        }
    },
    "1/4": {
        "major_diameter": 0.2500,
        "threads": {
            20: {
                "minor_diameter": 0.1887,
                "tap_75": 7,
                "tap_50": Fraction(7, 32),
            },
            28: {
                "minor_diameter": 0.2062,
                "tap_75": 3,
                "tap_50": 1,
            },
            32: {
                "minor_diameter": 0.2117,
                "tap_75": Fraction(7, 32),
                "tap_50": 1,
            }
        },
        "clearance": {"close_fit": 'F', "free_fit": 'H'},
        "shcs": {
            "hex": Fraction(3, 16),
            "counterbore_drill": Fraction(7, 16),
            "counterbore_dia": 0.4375,
            "counterbore_depth": 0.278
        }
    },
    "5/16": {
        "major_diameter": 0.3125,
        "threads": {
            18: {
                "minor_diameter": 0.2443,
                "tap_75": 'F',
                "tap_50": 'J',
            },
            24: {
                "minor_diameter": 0.2614,
                "tap_75": 'I',
                "tap_50": Fraction(9, 32),
            },
            32: {
                "minor_diameter": 0.2742,
                "tap_75": Fraction(9, 32),
                "tap_50": 'L',
            }
        },
        "clearance": {"close_fit": 'P', "free_fit": 'Q'},
        "shcs": {
            "hex": Fraction(1, 4),
            "counterbore_drill": Fraction(17, 32),
            "counterbore_dia": 0.53125,
            "counterbore_depth": 0.346
        }
    },
    "3/8": {
        "major_diameter": 0.3750,
        "threads": {
            16: {
                "minor_diameter": 0.2983,
                "tap_75": Fraction(5, 16),
                "tap_50": 'Q',
            },
            24: {
                "minor_diameter": 0.3239,
                "tap_75": 'Q',
                "tap_50": 'S',
            },
            32: {
                "minor_diameter": 0.3367,
                "tap_75": Fraction(11, 32),
                "tap_50": 'T',
            }
        },
        "clearance": {"close_fit": 'W', "free_fit": 'X'},
        "shcs": {
            "hex": Fraction(5, 16),
            "counterbore_drill": Fraction(5, 8),
            "counterbore_dia": 0.625,
            "counterbore_depth": 0.415
        }
    },
    "7/16": {
        "major_diameter": 0.4375,
        "threads": {
            14: {
                "minor_diameter": 0.3499,
                "tap_75": 'U',
                "tap_50": Fraction(25, 64),
            },
            20: {
                "minor_diameter": 0.3762,
                "tap_75": Fraction(25, 64),
                "tap_50": Fraction(13, 32),
            },
            28: {
                "minor_diameter": 0.3937,
                "tap_75": 'Y',
                "tap_50": 'Z',
            }
        },
        "clearance": {"close_fit": Fraction(29, 64), "free_fit": Fraction(15, 32)},
        "shcs": {
            "hex": Fraction(3, 8),
            "counterbore_drill": Fraction(23, 32),
            "counterbore_dia": 0.71875,
            "counterbore_depth": 0.483
        }
    },
    "1/2": {
        "major_diameter": 0.5000,
        "threads": {
            13: {
                "minor_diameter": 0.4056,
                "tap_75": Fraction(27, 64),
                "tap_50": Fraction(29, 64),
            },
            20: {
                "minor_diameter": 0.4387,
                "tap_75": Fraction(29, 64),
                "tap_50": Fraction(15, 32),
            },
            28: {
                "minor_diameter": 0.4562,
                "tap_75": Fraction(15, 32),
                "tap_50": Fraction(15, 32),
            }
        },
        "clearance": {"close_fit": Fraction(33, 64), "free_fit": Fraction(17, 32)},
        "shcs": {
            "hex": Fraction(3, 8),
            "counterbore_drill": Fraction(13, 16),
            "counterbore_dia": 0.8125,
            "counterbore_depth": 0.552
        }
    },
    "9/16": {
        "major_diameter": 0.5625,
        "threads": {
            12: {
                "minor_diameter": 0.4603,
                "tap_75": Fraction(31, 64),
                "tap_50": Fraction(33, 64),
            },
            18: {
                "minor_diameter": 0.4943,
                "tap_75": Fraction(33, 64),
                "tap_50": Fraction(17, 32),
            },
            24: {
                "minor_diameter": 0.5114,
                "tap_75": Fraction(33, 64),
                "tap_50": Fraction(17, 32),
            }
        },
        "clearance": {"close_fit": Fraction(37, 64), "free_fit": Fraction(19, 32)},
        "shcs": {
            "hex": Fraction(1, 2),
            "counterbore_drill": Fraction(29, 32),
            "counterbore_dia": 0.9062,
            "counterbore_depth": 0.594
        }
    },
    "5/8": {
        "major_diameter": 0.6250,
        "threads": {
            11: {
                "minor_diameter": 0.5135,
                "tap_75": Fraction(17, 32),
                "tap_50": Fraction(9, 16),
            },
            18: {
                "minor_diameter": 0.5568,
                "tap_75": Fraction(37, 64),
                "tap_50": Fraction(19, 32),
            },
            24: {
                "minor_diameter": 0.5739,
                "tap_75": Fraction(37, 64),
                "tap_50": Fraction(19, 32),
            }
        },
        "clearance": {"close_fit": Fraction(41, 64), "free_fit": Fraction(21, 32)},
        "shcs": {
            "hex": Fraction(1, 2),
            "counterbore_drill": Fraction(1, 1),
            "counterbore_dia": 1.0,
            "counterbore_depth": 0.689
        }
    },
    "11/16": {
        "major_diameter": 0.6875,
        "threads": {
            24: {
                "minor_diameter": 0.6364,
                "tap_75": Fraction(41, 64),
                "tap_50": Fraction(21, 32),
            }
        },
        "clearance": {"close_fit": Fraction(45, 64), "free_fit": Fraction(23, 32)}
    },
    "3/4": {
        "major_diameter": 0.7500,
        "threads": {
            10: {
                "minor_diameter": 0.6273,
                "tap_75": Fraction(21, 32),
                "tap_50": Fraction(11, 16),
            },
            16: {
                "minor_diameter": 0.6733,
                "tap_75": Fraction(11, 16),
                "tap_50": Fraction(45, 64),
            },
            20: {
                "minor_diameter": 0.6887,
                "tap_75": Fraction(45, 64),
                "tap_50": Fraction(23, 32),
            }
        },
        "clearance": {"close_fit": Fraction(49, 64), "free_fit": Fraction(25, 32)},
        "shcs": {
            "hex": Fraction(5, 8),
            "counterbore_drill": Fraction(19, 16),
            "counterbore_dia": 1.1875,
            "counterbore_depth": 0.828
        }
    },
    "13/16": {
        "major_diameter": 0.8125,
        "threads": {
            20: {
                "minor_diameter": 0.7512,
                "tap_75": Fraction(49, 64),
                "tap_50": Fraction(25, 32),
            }
        },
        "clearance": {"close_fit": Fraction(53, 64), "free_fit": Fraction(27, 32)}
    },
    "7/8": {
        "major_diameter": 0.8750,
        "threads": {
            9: {
                "minor_diameter": 0.7387,
                "tap_75": Fraction(49, 64),
                "tap_50": Fraction(51, 64),
            },
            14: {
                "minor_diameter": 0.7874,
                "tap_75": Fraction(13, 16),
                "tap_50": Fraction(53, 64),
            },
            20: {
                "minor_diameter": 0.8137,
                "tap_75": Fraction(53, 64),
                "tap_50": Fraction(27, 32),
            }
        },
        "clearance": {"close_fit": Fraction(57, 64), "free_fit": Fraction(29, 32)},
        "shcs": {
            "hex": Fraction(3, 4),
            "counterbore_drill": Fraction(11, 8),
            "counterbore_dia": 1.375,
            "counterbore_depth": 0.963
        }
    },
    "15/16": {
        "major_diameter": 0.9375,
        "threads": {
            20: {
                "minor_diameter": 0.8762,
                "tap_75": Fraction(57, 64),
                "tap_50": Fraction(29, 32),
            }
        },
        "clearance": {"close_fit": Fraction(61, 64), "free_fit": Fraction(31, 32)}
    },
    "1\"": {  # 1 inch screw
        "major_diameter": 1.0000,
        "threads": {
            8: {
                "minor_diameter": 0.8466,
                "tap_75": Fraction(7, 8),
                "tap_50": Fraction(59, 64),
            },
            12: {
                "minor_diameter": 0.8978,
                "tap_75": Fraction(15, 16),
                "tap_50": Fraction(61, 64),
            },
            20: {
                "minor_diameter": 0.9387,
                "tap_75": Fraction(61, 64),
                "tap_50": Fraction(31, 32),
            }
        },
        "clearance": {"close_fit": Fraction(65, 64), "free_fit": Fraction(33, 32)},
        "shcs": {
            "hex": Fraction(3, 4),
            "counterbore_drill": Fraction(13, 8),
            "counterbore_dia": 1.625,
            "counterbore_depth": 1.100
        }
    },
}

# ============================================================================
# SPREADSHEET GENERATION
# ============================================================================

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Tap & Drill Sizes"

# Define styles
header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
subheader_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
alt_row_fill = PatternFill(start_color="E2E2E2", end_color="E2E2E2", fill_type="solid")
bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center")
center_align_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Create header rows
# Row 1: Main headers with column spans
ws.merge_cells('A1:A2')  # Screw Size
ws['A1'] = "Screw Size"
ws['A1'].font = bold_font
ws['A1'].alignment = center_align_wrap

ws.merge_cells('B1:B2')  # Major Diameter
ws['B1'] = "Major Diameter"
ws['B1'].font = bold_font
ws['B1'].alignment = center_align_wrap

ws.merge_cells('C1:C2')  # Threads Per Inch
ws['C1'] = "TPI"
ws['C1'].font = bold_font
ws['C1'].alignment = center_align_wrap

ws.merge_cells('D1:D2')  # Minor Diameter
ws['D1'] = "Minor Diameter"
ws['D1'].font = bold_font
ws['D1'].alignment = center_align_wrap

# Tap Drill section
ws.merge_cells('E1:H1')  # Tap Drill header
ws['E1'] = "Tap Drill"
ws['E1'].font = bold_font
ws['E1'].alignment = center_align_wrap

ws.merge_cells('E2:F2')  # 75% Thread subsection
ws['E2'] = "75% Thread for Aluminum, Brass, Plastics"
ws['E2'].font = bold_font
ws['E2'].alignment = center_align_wrap

ws.merge_cells('G2:H2')  # 50% Thread subsection
ws['G2'] = "50% Thread for Stainless, Cast Iron & Iron"
ws['G2'].font = bold_font
ws['G2'].alignment = center_align_wrap

# Clearance Drill section
ws.merge_cells('I1:L1')  # Clearance Drill header
ws['I1'] = "Clearance Drill"
ws['I1'].font = bold_font
ws['I1'].alignment = center_align_wrap

ws.merge_cells('I2:J2')  # Close Fit subsection
ws['I2'] = "Close Fit"
ws['I2'].font = bold_font
ws['I2'].alignment = center_align_wrap

ws.merge_cells('K2:L2')  # Free Fit subsection
ws['K2'] = "Free Fit"
ws['K2'].font = bold_font
ws['K2'].alignment = center_align_wrap

# SHCS section
ws.merge_cells('M1:P1')  # SHCS header
ws['M1'] = "SHCS"
ws['M1'].font = bold_font
ws['M1'].alignment = center_align_wrap

ws.merge_cells('M2:M3')  # Hex subsection
ws['M2'] = "Hex"
ws['M2'].font = bold_font
ws['M2'].alignment = center_align_wrap

ws.merge_cells('N2:P2')  # Counterbore subsection
ws['N2'] = "Counterbore"
ws['N2'].font = bold_font
ws['N2'].alignment = center_align_wrap

# Row 3: Column detail headers
headers_row3 = [
    "", "", "", "",  # A-D (already merged from rows 1-2)
    "Drill Size", "Dec. Eq.",  # E-F (75% Thread)
    "Drill Size", "Dec. Eq.",  # G-H (50% Thread)
    "Drill Size", "Dec. Eq.",  # I-J (Close Fit)
    "Drill Size", "Dec. Eq.",  # K-L (Free Fit)
    "",  # M (Hex - merged from row 2)
    "Drill", "Dia.", "Depth"  # N-P (Counterbore)
]

for col, header in enumerate(headers_row3, start=1):
    if header:  # Skip empty cells (A-D)
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = bold_font
        cell.alignment = center_align_wrap

# Generate data rows from structured data
current_row = 4
alternate_color = False  # Track alternating colors for screw sizes

# Sort screw sizes by major diameter
sorted_screws = sorted(THREAD_DATA.items(), key=lambda x: x[1]['major_diameter'])

for screw_size, screw_data in sorted_screws:
    major_diam = screw_data["major_diameter"]
    threads = screw_data["threads"]
    clearance = screw_data["clearance"]
    
    # Toggle alternating color for each new screw size
    alternate_color = not alternate_color
    
    # Get thread pitches sorted (ascending order - coarse threads first)
    tpis = sorted(threads.keys())
    num_threads = len(tpis)
    
    # Process each thread pitch for this screw size
    first_row_of_screw = current_row
    for idx, tpi in enumerate(tpis):
        thread_spec = threads[tpi]
        
        # Determine if we need to merge cells
        is_first_thread = (idx == 0)
        
        # Write screw size and major diameter (only on first thread row)
        if is_first_thread:
            ws.cell(row=current_row, column=1).value = format_screw_size(screw_size)
            ws.cell(row=current_row, column=2).value = format_decimal(major_diam)
            
            # Merge screw size and major diameter if multiple threads
            if num_threads > 1:
                ws.merge_cells(f'A{first_row_of_screw}:A{first_row_of_screw + num_threads - 1}')
                ws.merge_cells(f'B{first_row_of_screw}:B{first_row_of_screw + num_threads - 1}')
        
        # Write TPI and minor diameter
        ws.cell(row=current_row, column=3).value = str(tpi)
        ws.cell(row=current_row, column=4).value = format_decimal(thread_spec['minor_diameter'])
        
        # Write tap drill sizes (75% thread)
        tap_75 = thread_spec['tap_75']
        ws.cell(row=current_row, column=5).value = format_drill_size(tap_75)
        ws.cell(row=current_row, column=6).value = format_decimal(get_drill_decimal(tap_75))
        
        # Write tap drill sizes (50% thread)
        tap_50 = thread_spec['tap_50']
        ws.cell(row=current_row, column=7).value = format_drill_size(tap_50)
        ws.cell(row=current_row, column=8).value = format_decimal(get_drill_decimal(tap_50))
        
        # Write clearance drill sizes (only on first thread row, then merge)
        if is_first_thread:
            close_fit = clearance['close_fit']
            ws.cell(row=current_row, column=9).value = format_drill_size(close_fit)
            ws.cell(row=current_row, column=10).value = format_decimal(get_drill_decimal(close_fit))
            
            free_fit = clearance['free_fit']
            ws.cell(row=current_row, column=11).value = format_drill_size(free_fit)
            ws.cell(row=current_row, column=12).value = format_decimal(get_drill_decimal(free_fit))
            
            # Merge clearance drill columns if multiple threads
            if num_threads > 1:
                ws.merge_cells(f'I{first_row_of_screw}:I{first_row_of_screw + num_threads - 1}')
                ws.merge_cells(f'J{first_row_of_screw}:J{first_row_of_screw + num_threads - 1}')
                ws.merge_cells(f'K{first_row_of_screw}:K{first_row_of_screw + num_threads - 1}')
                ws.merge_cells(f'L{first_row_of_screw}:L{first_row_of_screw + num_threads - 1}')
            
            # Write SHCS data (only on first thread row, then merge)
            if 'shcs' in screw_data:
                shcs = screw_data['shcs']
                ws.cell(row=current_row, column=13).value = format_drill_size(shcs['hex'])
                ws.cell(row=current_row, column=14).value = format_drill_size(shcs['counterbore_drill'])
                ws.cell(row=current_row, column=15).value = format_decimal(shcs['counterbore_dia'])
                ws.cell(row=current_row, column=16).value = format_decimal(shcs['counterbore_depth'])
                
                # Merge SHCS columns if multiple threads
                if num_threads > 1:
                    ws.merge_cells(f'M{first_row_of_screw}:M{first_row_of_screw + num_threads - 1}')
                    ws.merge_cells(f'N{first_row_of_screw}:N{first_row_of_screw + num_threads - 1}')
                    ws.merge_cells(f'O{first_row_of_screw}:O{first_row_of_screw + num_threads - 1}')
                    ws.merge_cells(f'P{first_row_of_screw}:P{first_row_of_screw + num_threads - 1}')
        
        # Apply formatting to all cells in this row
        for col in range(1, 17):
            cell = ws.cell(row=current_row, column=col)
            if not isinstance(cell, MergedCell):
                cell.alignment = center_align
                cell.border = thin_border
                # Apply alternating background color
                if alternate_color:
                    cell.fill = alt_row_fill
        
        current_row += 1

# Apply borders to all header cells
for row in range(1, 4):
    for col in range(1, 17):
        ws.cell(row=row, column=col).border = thin_border

# Apply borders to all data cells (including merged cells in the last row)
last_data_row = current_row - 1
for row in range(4, last_data_row + 1):
    for col in range(1, 17):
        cell = ws.cell(row=row, column=col)
        # Apply border even to merged cells to ensure bottom borders appear
        cell.border = thin_border

# Adjust column widths
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
    ws.column_dimensions[col].width = 11
for col in ['M', 'N', 'O', 'P']:
    ws.column_dimensions[col].width = 11

# Set row heights for better readability
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 40
ws.row_dimensions[3].height = 20

# Configure page setup for printing
# Page size: 1=Letter, 3=Tabloid (11"×17"), 5=Legal, 9=A4
ws.page_setup.paperSize = 3  # Tabloid size (11×17)
ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToPage = True
ws.page_setup.fitToHeight = 1  # Fit to 1 page tall
ws.page_setup.fitToWidth = 1   # Fit to 1 page wide

# Print options
ws.print_options.horizontalCentered = True
ws.print_options.verticalCentered = False
ws.print_options.gridLines = False

# Page margins (in inches) - 0.5" all around
ws.page_margins.left = 0.5
ws.page_margins.right = 0.5
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5
ws.page_margins.header = 0.0  # No header
ws.page_margins.footer = 0.0  # No footer

# Save as Excel format (LibreOffice can open this)
wb.save('inch_taps_drills.xlsx')
print("Spreadsheet created: inch_taps_drills.xlsx")
print("This file can be opened in LibreOffice Calc with all merged cells preserved.")
