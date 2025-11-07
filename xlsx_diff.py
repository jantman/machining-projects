#!/usr/bin/env python3
"""
Compare two Excel (.xlsx) files and show differences.

Usage:
    python xlsx_diff.py file1.xlsx file2.xlsx
    python xlsx_diff.py file1.xlsx file2.xlsx --sheet "Sheet Name"
"""

import sys
import openpyxl
from openpyxl.cell.cell import MergedCell

def get_cell_value(cell):
    """Get the value from a cell, handling merged cells."""
    if isinstance(cell, MergedCell):
        return None
    return cell.value

def compare_sheets(ws1, ws2, sheet_name):
    """Compare two worksheets and return differences."""
    differences = []
    
    # Get the dimensions
    max_row = max(ws1.max_row, ws2.max_row)
    max_col = max(ws1.max_column, ws2.max_column)
    
    print(f"\nComparing sheet: {sheet_name}")
    print(f"  File 1: {ws1.max_row} rows × {ws1.max_column} columns")
    print(f"  File 2: {ws2.max_row} rows × {ws2.max_column} columns")
    
    diff_count = 0
    
    # Compare cell by cell
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell1 = ws1.cell(row=row, column=col) if row <= ws1.max_row and col <= ws1.max_column else None
            cell2 = ws2.cell(row=row, column=col) if row <= ws2.max_row and col <= ws2.max_column else None
            
            val1 = get_cell_value(cell1) if cell1 else None
            val2 = get_cell_value(cell2) if cell2 else None
            
            # Skip if both are None or MergedCell
            if val1 is None and val2 is None:
                continue
            
            if val1 != val2:
                col_letter = openpyxl.utils.get_column_letter(col)
                differences.append({
                    'cell': f"{col_letter}{row}",
                    'row': row,
                    'col': col,
                    'file1': val1,
                    'file2': val2
                })
                diff_count += 1
    
    return differences, diff_count

def print_differences(differences, limit=50):
    """Print the differences found."""
    if not differences:
        print("  ✓ No differences found!")
        return
    
    print(f"\n  Found {len(differences)} differences:")
    
    for i, diff in enumerate(differences[:limit]):
        print(f"\n  Cell {diff['cell']} (Row {diff['row']}, Col {diff['col']}):")
        print(f"    File 1: {repr(diff['file1'])}")
        print(f"    File 2: {repr(diff['file2'])}")
    
    if len(differences) > limit:
        print(f"\n  ... and {len(differences) - limit} more differences (showing first {limit})")

def compare_workbooks(file1, file2, sheet_name=None):
    """Compare two Excel workbooks."""
    print(f"Comparing:")
    print(f"  File 1: {file1}")
    print(f"  File 2: {file2}")
    
    try:
        wb1 = openpyxl.load_workbook(file1, data_only=True)
        wb2 = openpyxl.load_workbook(file2, data_only=True)
    except Exception as e:
        print(f"Error loading files: {e}")
        return False
    
    # Compare sheet names
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    
    print(f"\nSheets in File 1: {sheets1}")
    print(f"Sheets in File 2: {sheets2}")
    
    if sheets1 != sheets2:
        print("\n⚠ Warning: Sheet names differ!")
        only_in_1 = set(sheets1) - set(sheets2)
        only_in_2 = set(sheets2) - set(sheets1)
        if only_in_1:
            print(f"  Only in File 1: {only_in_1}")
        if only_in_2:
            print(f"  Only in File 2: {only_in_2}")
    
    # Compare sheets
    if sheet_name:
        if sheet_name not in sheets1:
            print(f"\nError: Sheet '{sheet_name}' not found in File 1")
            return False
        if sheet_name not in sheets2:
            print(f"\nError: Sheet '{sheet_name}' not found in File 2")
            return False
        sheets_to_compare = [sheet_name]
    else:
        sheets_to_compare = set(sheets1) & set(sheets2)
    
    total_differences = 0
    for sheet in sheets_to_compare:
        ws1 = wb1[sheet]
        ws2 = wb2[sheet]
        differences, diff_count = compare_sheets(ws1, ws2, sheet)
        total_differences += diff_count
        print_differences(differences)
    
    print(f"\n{'='*60}")
    if total_differences == 0:
        print("✓ Files are identical!")
    else:
        print(f"✗ Found {total_differences} total differences")
    print('='*60)
    
    return total_differences == 0

def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    
    file1 = sys.argv[1]
    file2 = sys.argv[2]
    
    sheet_name = None
    if len(sys.argv) > 3 and sys.argv[3] == '--sheet' and len(sys.argv) > 4:
        sheet_name = sys.argv[4]
    
    try:
        are_identical = compare_workbooks(file1, file2, sheet_name)
        sys.exit(0 if are_identical else 1)
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(2)

if __name__ == '__main__':
    main()
